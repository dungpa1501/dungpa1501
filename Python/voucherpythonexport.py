import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# SQL Server connection details
server = '192.168.115.3'
database = 'VISTAVM' 
username = 'sa'
password = 'GaLaXyNDNTTB126'
driver = '{ODBC Driver 17 for SQL Server}'  # Ensure you have the correct driver installed

# Create a connection string
connection_string = f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}'

# Connect to SQL Server
conn = pyodbc.connect(connection_string, timeout=120)

# variables for date
from_date = 'DATEADD(MONTH, DATEDIFF(MONTH, 0, GETDATE())-1, 0)'
to_date = 'DATEADD(MONTH, DATEDIFF(MONTH, 0, GETDATE()), 0)'

# SQL Query
query_tondau = f'''
    declare @fromdate datetime, @todate datetime
    set @fromdate = {from_date}
    set @todate = {to_date}

    select memberId,VStock_strBookletIdent,e.sName StockLocation,f.sName IssuedLocation,Stock_strBarcode, nVoucherCode, Stock_dtmCreated ,dIssuedDate, dExpiryDate, dRedeemedDate, 
    case when dRedeemedDate is not null then 'Redeemed' 
        when dExpiryDate < cast(cast(getdate() as date) as datetime) then 'Expired'
        else 'Usable' end status
    from tblStock a 
    left join tblRedeemed b on a.lVoucherTypeID = b.lVoucherTypeID and a.lVoucherNumber = b.lVoucherNumber and a.nDuplicateNo = b.nDuplicateNo
    left join glxMemberVoucher c on a.Stock_strBarcode = c.voucherCode
    join tblVoucherType d on a.lVoucherTypeID = d.lID
	left join tblLocation e on a.lStockLocationID = e.lID
	left join tblLocation f on a.lIssuedLocationID = f.lID				    
    where 1=1
    and Stock_dtmCreated < @fromdate
    and dRedeemedDate is null 
    and year(dIssuedDate) = 1900
    and (dExpiryDate >=  cast(cast(getdate() as date) as datetime) or dExpiryDate is null)
'''

query_nhap = f'''
    declare @fromdate datetime, @todate datetime
    set @fromdate = {from_date}
    set @todate = {to_date}

    select memberId,VStock_strBookletIdent,e.sName StockLocation,f.sName IssuedLocation,Stock_strBarcode, nVoucherCode, Stock_dtmCreated ,dIssuedDate, dExpiryDate, dRedeemedDate, 
    case when dRedeemedDate is not null then 'Redeemed' 
        when dExpiryDate < cast(cast(getdate() as date) as datetime) then 'Expired'
        else 'Usable' end status
    from tblStock a 
    left join tblRedeemed b on a.lVoucherTypeID = b.lVoucherTypeID and a.lVoucherNumber = b.lVoucherNumber and a.nDuplicateNo = b.nDuplicateNo
    left join glxMemberVoucher c on a.Stock_strBarcode = c.voucherCode
    join tblVoucherType d on a.lVoucherTypeID = d.lID
	left join tblLocation e on a.lStockLocationID = e.lID
	left join tblLocation f on a.lIssuedLocationID = f.lID    
    where 1=1
    and Stock_dtmCreated >= @fromdate
    and Stock_dtmCreated < @todate
'''

query_xuat = f'''
    declare @fromdate datetime, @todate datetime
    set @fromdate = {from_date}
    set @todate = {to_date}

    select memberId,VStock_strBookletIdent,e.sName StockLocation,f.sName IssuedLocation,Stock_strBarcode, nVoucherCode, Stock_dtmCreated ,dIssuedDate, dExpiryDate, dRedeemedDate, 
    case when dRedeemedDate is not null then 'Redeemed' 
        when dExpiryDate < cast(cast(getdate() as date) as datetime) then 'Expired'
        else 'Usable' end status
    from tblStock a 
    left join tblRedeemed b on a.lVoucherTypeID = b.lVoucherTypeID and a.lVoucherNumber = b.lVoucherNumber and a.nDuplicateNo = b.nDuplicateNo
    left join glxMemberVoucher c on a.Stock_strBarcode = c.voucherCode
    join tblVoucherType d on a.lVoucherTypeID = d.lID
	left join tblLocation e on a.lStockLocationID = e.lID
	left join tblLocation f on a.lIssuedLocationID = f.lID    
    where 1=1
    and dIssuedDate >= @fromdate
    and dIssuedDate < @todate
'''

query_toncuoi = f'''
    declare @fromdate datetime, @todate datetime
    set @fromdate = {from_date}
    set @todate = {to_date}

    select memberId,VStock_strBookletIdent,e.sName StockLocation,f.sName IssuedLocation,Stock_strBarcode, nVoucherCode, Stock_dtmCreated ,dIssuedDate, dExpiryDate, dRedeemedDate, 
    case when dRedeemedDate is not null then 'Redeemed' 
        when dExpiryDate < cast(cast(getdate() as date) as datetime) then 'Expired'
        else 'Usable' end status
    from tblStock a 
    left join tblRedeemed b on a.lVoucherTypeID = b.lVoucherTypeID and a.lVoucherNumber = b.lVoucherNumber and a.nDuplicateNo = b.nDuplicateNo
    left join glxMemberVoucher c on a.Stock_strBarcode = c.voucherCode
    join tblVoucherType d on a.lVoucherTypeID = d.lID
	left join tblLocation e on a.lStockLocationID = e.lID
	left join tblLocation f on a.lIssuedLocationID = f.lID    
    where 1=1
    --and Stock_dtmCreated >= '2024-08-01 00:00:00.000'
    and Stock_dtmCreated < @todate
    and year(dIssuedDate) = 1900
    and dRedeemedDate is null 
    and (dExpiryDate >= cast(cast(getdate() as date) as datetime) or dExpiryDate is null)
'''


# Use pandas to execute the query and read the data into a DataFrame
df_tondau = pd.read_sql(query_tondau, conn)
df_nhap = pd.read_sql(query_nhap, conn)
df_xuat = pd.read_sql(query_xuat, conn)
df_toncuoi = pd.read_sql(query_toncuoi, conn)

# Export the DataFrame to Excel
with pd.ExcelWriter(r'E:\voucher_data.xlsx', engine='openpyxl') as writer:
    df_tondau.to_excel(writer, index=False, sheet_name='Tồn Đầu')
    df_nhap.to_excel(writer,index=False,sheet_name='Nhập')
    df_xuat.to_excel(writer,index=False,sheet_name='Xuất')
    df_toncuoi.to_excel(writer,index=False,sheet_name='Tồn cuối')

    # Access the workbook and the sheets for formatting
    workbook = writer.book
    sheet_tondau = workbook['Tồn Đầu']
    sheet_nhap = workbook['Nhập']
    sheet_xuat = workbook['Xuất']
    sheet_toncuoi = workbook['Tồn cuối']

    # Style settings: Light blue fill for headers and square borders
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply header style and auto-resize columns for the "tondau" sheet
    for col in sheet_tondau.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get column letter
        for cell in col:
            # Auto-adjust column width
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            cell.border = thin_border  # Apply border to every cell
        sheet_tondau.column_dimensions[col_letter].width = max_length + 2  # Add padding

    # Apply header style for the "tondau" sheet
    for cell in sheet_tondau[1]:
        cell.fill = header_fill  # Light blue header background
        cell.border = thin_border

    # Apply header style and auto-resize columns for the "nhap" sheet
    for col in sheet_nhap.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get column letter
        for cell in col:
            # Auto-adjust column width
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            cell.border = thin_border  # Apply border to every cell
        sheet_nhap.column_dimensions[col_letter].width = max_length + 2  # Add padding

    # Apply header style for the "nhap" sheet
    for cell in sheet_nhap[1]:
        cell.fill = header_fill  # Light blue header background
        cell.border = thin_border

    # Apply header style and auto-resize columns for the "xuat" sheet
    for col in sheet_xuat.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get column letter
        for cell in col:
            # Auto-adjust column width
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            cell.border = thin_border  # Apply border to every cell
        sheet_xuat.column_dimensions[col_letter].width = max_length + 2  # Add padding

    # Apply header style for the "xuat" sheet
    for cell in sheet_xuat[1]:
        cell.fill = header_fill  # Light blue header background
        cell.border = thin_border
    
        # Apply header style and auto-resize columns for the "toncuoi" sheet
    for col in sheet_toncuoi.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get column letter
        for cell in col:
            # Auto-adjust column width
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            cell.border = thin_border  # Apply border to every cell
        sheet_toncuoi.column_dimensions[col_letter].width = max_length + 2  # Add padding

    # Apply header style for the "xuat" sheet
    for cell in sheet_toncuoi[1]:
        cell.fill = header_fill  # Light blue header background
        cell.border = thin_border    

# Close the connection
conn.close()

print("Data exported successfully to voucher_datas.xlsx")
